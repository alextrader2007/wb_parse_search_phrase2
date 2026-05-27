#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
Парсер товаров Wildberries (WB) — сбор данных через публичные API.

Что умеет:
  1) Получать информацию по списку артикулов (SKU): цены, остатки, склады
  2) Искать товары по ключевому слову и определять их позицию в выдаче
  3) Работать через прокси (чтобы не заблокировали)
  4) Сохранять результат в Excel (.xlsx) с авто-фильтром, авто-шириной колонок и картинками
  5) Выгружать характеристики и описание каждого товара на отдельные листы
  6) Опционально сохранять CSV (по запросу пользователя)

Как устроен парсер (для новичков):
  - Весь код — это последовательность функций: одна вызывает другую
  - main() — точка входа (запускается при старте)
  - main() спрашивает у пользователя, что делать, и вызывает нужные функции
  - Каждая функция делает только свою работу и возвращает результат

Внутреннее устройство (защита WB):
  - WB (Wildberries) требует специальный токен x_wbaas_token для внутренних API
  - Токен выдаётся после прохождения JS-челленджа (проверки, что вы не бот)
  - Парсер получает токен через SeleniumBase (настоящий браузер Chrome)
  - Если SeleniumBase не установлен — пробует curl_cffi, затем обычный HTTP
  - Токен живёт ~неделю, кэшируется в переменной _WBAAS_TOKEN на время сеанса

Особенности сборки .exe (PyInstaller):
  - Драйвер Chrome сохраняется в %LOCALAPPDATA%\\WBParser\\drivers (не в temp)
  - Pillow включается через --collect-all PIL (нужен для вставки изображений в Excel)
  - После завершения — Enter для повтора, Esc для выхода

Известные ограничения:
  - totalQuantity: 39 — заглушка WB для «в наличии, точное количество скрыто».
    Парсер игнорирует эту заглушку и показывает только сумму stocks[].qty.
  - Цены всегда запрашиваются через PRIMARY_DEST (dest=-2888067, BYN).
    Остальные регионы (RUB, KZT) используются только для stocks[].
"""

import sys
import io
import os
import time
import random
import argparse
import requests          # библиотека для отправки HTTP-запросов (основная работа с сетью)
import pandas as pd      # библиотека для работы с таблицами (Excel/CSV)
import msvcrt            # чтение клавиш (Enter/Esc) без ожидания Enter
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime
from typing import List, Dict, Any, Optional
from openpyxl.styles import Font, Alignment
from decimal import Decimal, ROUND_FLOOR  # точные математические расчёты (цены с WB-кошельком)

# ---------------------------------------------------------------------------
# Настройка кодировки терминала Windows (чтобы кириллица отображалась корректно)
# ---------------------------------------------------------------------------
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

# ---------------------------------------------------------------------------
# Rich — библиотека для красивых цветных таблиц и прогресс-баров в консоли
# Если не установлена — используется заглушка (простой print)
# ---------------------------------------------------------------------------
try:
    from rich.console import Console
    from rich.table import Table
    from rich.progress import Progress, SpinnerColumn, TextColumn, BarColumn, TimeRemainingColumn
    from rich.panel import Panel
    from rich.prompt import Prompt, Confirm
    from rich import box
except ImportError:
    class Console:
        def print(self, *args, **kwargs):
            print(*args, **kwargs)
        def rule(self, *args, **kwargs):
            print("-" * 50)
    Console = Console()

console = Console()

# ---------------------------------------------------------------------------
# ═══════════════════════════════════════════════════════════════════════════
# КОНСТАНТЫ (настройки, которые можно менять)
# ═══════════════════════════════════════════════════════════════════════════
# ---------------------------------------------------------------------------

# ID региона для получения цен в BYN (Беларусь, Гродно).
# У Wildberries цены и остатки товаров зависят от региона (dest).
# Если нужны цены в рублях (RUB) — замени на "-1257786" (Москва).
DEFAULT_DEST = "-2888067"

# ID Москвы — используется для второго запроса, чтобы получить остатки по ВСЕМ складам.
# Проблема: WB для регионов Беларуси не отдаёт остатки по складам (только "ближайший склад").
# Москва отдаёт полную картину, поэтому делаем два запроса на каждый товар.
MOSCOW_DEST = "-1257786"

# Валюта, в которой возвращаются цены
DEFAULT_CURR = "byn"  # byn = белорусские рубли, rub = российские рубли

# HTTP-заголовки, которые парсер отправляет вместе с каждым запросом.
# User-Agent имитирует обычный браузер Chrome, чтобы WB не заподозрил бота.
DEFAULT_HEADERS = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
    'Accept': '*/*',
    'Accept-Language': 'ru-RU,ru;q=0.9,en-US;q=0.8,en;q=0.7',
    'Origin': 'https://www.wildberries.ru',
    'Referer': 'https://www.wildberries.ru/'
}

# ---------------------------------------------------------------------------
# ГЛОБАЛЬНЫЕ ПЕРЕМЕННЫЕ (живут всё время работы программы)
# ---------------------------------------------------------------------------

_WBAAS_TOKEN = None      # Кэш для x_wbaas_token — чтобы не получать его заново для каждого запроса
_SESSION = requests.Session()  # Сессия HTTP — сохраняет соединения (Keep-Alive), ускоряет запросы
_SESSION.headers.update(DEFAULT_HEADERS)

_VOL_BASKET_CACHE = {}   # Кэш для корзин изображений — чтобы не перебирать basket-* каждый раз
_BASKET_EXECUTOR = None  # Пул потоков для параллельной проверки корзин

# ───────────────────────────────────────────────────────────────────────────
# ВСПОМОГАТЕЛЬНАЯ ФУНКЦИЯ: статическое определение корзины (basket) по vol
# ───────────────────────────────────────────────────────────────────────────

def _get_basket_static(vol: int) -> str:
    if vol <= 143: return "01"
    elif vol <= 287: return "02"
    elif vol <= 431: return "03"
    elif vol <= 719: return "04"
    elif vol <= 1007: return "05"
    elif vol <= 1061: return "06"
    elif vol <= 1115: return "07"
    elif vol <= 1169: return "08"
    elif vol <= 1313: return "09"
    elif vol <= 1601: return "10"
    elif vol <= 1655: return "11"
    elif vol <= 1919: return "12"
    elif vol <= 2045: return "13"
    elif vol <= 2189: return "14"
    elif vol <= 2405: return "15"
    return f"{(16 + (vol - 2406) // 216):02d}"

# ───────────────────────────────────────────────────────────────────────────
# ВСПОМОГАТЕЛЬНАЯ ФУНКЦИЯ: определение корзины (basket) для изображений
# ───────────────────────────────────────────────────────────────────────────

def get_basket_dynamically(vol: int, part: int, sku: int) -> str:
    """
    Определяет номер корзины (basket-XX) для загрузки изображений товара.

    Зачем это нужно:
      - Изображения товаров WB лежат на CDN: basket-XX.wbbasket.ru/vol{vol}/...
      - Номер корзины (XX) зависит от "vol" (объёма) товара.
      - Раньше номера корзин были жёстко привязаны к vol, но WB изменил логику.
      - Теперь приходится проверять несколько корзин, пока не найдём нужную.

    Как работает:
      1) Сначала угадываем корзину по статической таблице (быстрая эвристика)
      2) Проверяем популярные корзины 39, 40, 41 (туда WB переносит много товаров)
      3) Если не нашли — параллельно проверяем корзины 01..99 (ThreadPool)

    Args:
        vol: объём товара (vol = артикул // 100000)
        part: часть (part = артикул // 1000)
        sku: артикул товара

    Returns:
        строка с номером корзины, например "01", "12", "39"
    """
    # Если уже знаем корзину для этого vol — возвращаем из кэша
    if vol in _VOL_BASKET_CACHE:
        return _VOL_BASKET_CACHE[vol]

    global _BASKET_EXECUTOR
    if _BASKET_EXECUTOR is None:
        _BASKET_EXECUTOR = ThreadPoolExecutor(max_workers=15)

    # ── Шаг 1: статическая эвристика (по старым правилам WB) ──
    guess = _get_basket_static(vol)

    def check_image(b_id_str):
        url = f"https://basket-{b_id_str}.wbbasket.ru/vol{vol}/part{part}/{sku}/images/big/1.webp"
        try:
            if _SESSION.head(url, timeout=1.5).status_code == 200:
                return b_id_str
        except Exception:
            pass
        return None

    # ── Шаг 2: проверяем "угаданную" корзину (быстрый путь) ──
    if check_image(guess):
        _VOL_BASKET_CACHE[vol] = guess
        return guess

    # ── Шаг 3: проверяем корзины 39, 40, 41 (туда WB массово переносит товары) ──
    for common_basket in ['39', '40', '41']:
        if common_basket != guess and check_image(common_basket):
            _VOL_BASKET_CACHE[vol] = common_basket
            return common_basket

    # ── Шаг 4: параллельная проверка всех корзин 01..99 ──
    baskets_to_check = [f"{i:02d}" for i in range(1, 201) if f"{i:02d}" not in [guess, '39', '40', '41']]

    futures = [_BASKET_EXECUTOR.submit(check_image, b) for b in baskets_to_check]
    for future in as_completed(futures):
        result = future.result()
        if result:
            _VOL_BASKET_CACHE[vol] = result
            return result

    # Если ничего не нашли — возвращаем угаданную (хотя бы попытка)
    _VOL_BASKET_CACHE[vol] = guess
    return guess


# ───────────────────────────────────────────────────────────────────────────
# ВСПОМОГАТЕЛЬНАЯ ФУНКЦИЯ: безопасный HTTP-запрос с повторными попытками
# ───────────────────────────────────────────────────────────────────────────

def make_request(url: str, params=None, headers: Optional[Dict[str, Any]] = None,
                 cookies: Optional[Dict[str, str]] = None,
                 proxies: Optional[List[str]] = None,
                 max_retries: int = 3, timeout: int = 15) -> requests.Response:
    """
    Универсальная функция для выполнения HTTP GET-запроса.

    Особенности:
      - При ответе 429 (Too Many Requests) ждёт и повторяет (exponential backoff)
      - При 404 сразу выходит с ошибкой (нет смысла повторять)
      - Использует глобальную сессию _SESSION для Keep-Alive

    Args:
        url: адрес запроса
        params: параметры запроса (dict или список кортежей)
        headers: HTTP-заголовки
        cookies: Cookie (например, x_wbaas_token)
        proxies: список прокси (выбирается случайный)
        max_retries: сколько раз повторять при ошибке
        timeout: таймаут ожидания ответа (секунд)

    Returns:
        объект ответа requests.Response

    Raises:
        requests.exceptions.HTTPError: если все попытки исчерпаны
    """
    delay = 3.0
    for attempt in range(max_retries):
        try:
            # Выбираем случайный прокси из списка (если есть)
            if proxies:
                selected = random.choice(proxies)
                proxy = {'http': selected, 'https': selected}
            else:
                proxy = None

            response = _SESSION.get(url, params=params, headers=headers,
                                    cookies=cookies, proxies=proxy, timeout=timeout)

            # WB при слишком частых запросах возвращает 429
            if response.status_code == 429:
                retry_after = response.headers.get("Retry-After")
                wait_time = int(retry_after) if (retry_after and retry_after.isdigit()) else int(delay)
                console.print(f"[yellow]\n[!] Предупреждение (429): Слишком много запросов. "
                              f"Ожидание {wait_time} сек. перед повторной попыткой...[/yellow]")
                time.sleep(wait_time)
                delay *= 2
                continue

            response.raise_for_status()  # выбросит исключение, если статус не 2xx
            return response

        except requests.exceptions.RequestException as e:
            # 404 — товар не найден, повторять бессмысленно
            if isinstance(e, requests.exceptions.HTTPError) and e.response is not None and e.response.status_code == 404:
                raise e
            if attempt == max_retries - 1:
                raise e  # последняя попытка — даём ошибке уйти наверх
            time.sleep(delay)
            delay *= 2

    raise requests.exceptions.HTTPError("Превышено количество попыток запроса из-за ограничений 429")


# ───────────────────────────────────────────────────────────────────────────
# ВСПОМОГАТЕЛЬНАЯ ФУНКЦИЯ: загрузка справочника складов WB
# ───────────────────────────────────────────────────────────────────────────

def fetch_warehouse_map(proxies: Optional[List[str]] = None) -> Dict[int, str]:
    """
    Загружает с CDN WB список складов и их названия.

    Зачем: в ответах API WB склады обозначаются числовыми ID (например, 117986).
    Чтобы показывать пользователю понятные названия ("Коледино", "Электросталь"),
    мы загружаем соответствие ID → имя.

    Returns:
        словарь {id_склада: "название склада", ...}
    """
    url = "https://static-basket-01.wbbasket.ru/vol0/data/stores-data.json"

    try:
        response = make_request(url, headers=DEFAULT_HEADERS, proxies=proxies, timeout=10)
        response.encoding = 'utf-8'
        stores_list = response.json()
        return {store['id']: store['name'] for store in stores_list if 'id' in store and 'name' in store}
    except Exception as e:
        console.print(f"[yellow]Предупреждение: Не удалось загрузить карту складов ({e}). "
                       f"ID складов будут отображаться без имён.[/yellow]")
    return {}


# ───────────────────────────────────────────────────────────────────────────
# ОСНОВНАЯ ФУНКЦИЯ ПАРСИНГА: превращает сырой JSON от WB в понятный словарь
# ───────────────────────────────────────────────────────────────────────────

def parse_single_product(product: Dict[str, Any],
                         warehouse_map: Dict[int, str],
                         search_data: Optional[Dict[str, Any]] = None) -> Dict[str, Any]:
    """
    Берёт "сырой" JSON товара от API WB и превращает в плоский словарь с понятными ключами.

    Что делает:
      1) Извлекает основные поля: артикул, название, бренд, продавец, рейтинг
      2) Достаёт цены (в копейках, делит на 100)
      3) Собирает остатки по складам и суммирует общее количество
      4) Формирует ссылки на товар и изображение
      5) Если есть search_data — добавляет позицию в выдаче и информацию о доставке

    Args:
        product: словарь товара из API WB (JSON, преобразованный в dict)
        warehouse_map: словарь {ID_склада: имя_склада}
        search_data: данные из поисковой выдачи (позиция, реклама, доставка)

    Returns:
        словарь с полями: Артикул, Бренд, Название, Цена без скидки,
        Цена со скидкой, Цена с WB кошельком, Рейтинг, Отзывы (кол-во),
        Остатки всего (шт), Детализация по складам, Ссылка на товар, Ссылка на фото
    """
    # ── Основные поля ──
    product_id = product.get('id')
    name = product.get('name', 'Неизвестно')
    brand = product.get('brand') or 'Без бренда'
    supplier = product.get('supplier', 'Неизвестный продавец')
    supplier_id = product.get('supplierId')
    rating = product.get('rating', 0)
    feedbacks = product.get('feedbacks', 0)

    # ── Цены ──
    # В API WB цены передаются в копейках (целые числа):
    #   priceU = цена без скидки (в копейках)
    #   salePriceU = цена со скидкой (в копейках)
    # Делим на 100, чтобы получить рубли/копейки.
    price_u = product.get('priceU')
    sale_price_u = product.get('salePriceU')

    # Иногда цена лежит внутри sizes[].price (на карточке товара через v4/detail),
    # а не в корне объекта. Подстраховываемся:
    sizes = product.get('sizes', [])
    if (price_u is None or sale_price_u is None) and sizes:
        first_size = sizes[0]
        price_obj = first_size.get('price', {})
        if price_obj:
            price_u = price_obj.get('basic')
            sale_price_u = price_obj.get('product')

    price_original = price_u / 100.0 if price_u else 0.0
    price_discounted = sale_price_u / 100.0 if sale_price_u else 0.0

    # ── Остатки по складам ──
    sizes_list = []
    stocks_detail = []
    total_stock = 0

    sizes = product.get('sizes', [])
    for size_obj in sizes:
        size_name = size_obj.get('origName') or size_obj.get('name') or 'No Size'
        sizes_list.append(size_name)

        stocks = size_obj.get('stocks', [])
        for stock_obj in stocks:
            wh_id = stock_obj.get('wh')
            qty = stock_obj.get('qty', 0)
            total_stock += qty
            wh_name = warehouse_map.get(wh_id, f"Склад ID {wh_id}")
            stocks_detail.append({
                'size': size_name,
                'wh_id': wh_id,
                'wh_name': wh_name,
                'qty': qty
            })

    sizes_str = ", ".join(map(str, sizes_list))

    # totalQuantity — это общее количество товара на ВСЕХ складах.
    # Оно может быть больше, чем сумма по размерам (если какие-то размеры не попали в ответ).
    total_quantity = product.get('totalQuantity')
    if total_quantity is not None and total_quantity not in (39, 0) and total_quantity > total_stock:
        total_stock = total_quantity

    # Группируем остатки по складам (суммируем все размеры в рамках одного склада)
    wh_agg = {}
    for item in stocks_detail:
        name_wh = item['wh_name']
        wh_agg[name_wh] = wh_agg.get(name_wh, 0) + item['qty']

    wh_summary_str = ", ".join([f"{wh}: {qty}" for wh, qty in wh_agg.items()]) if wh_agg else "Нет в наличии"

    # ── Ссылка на изображение ──
    image_url = ""
    if product_id:
        vol = product_id // 100000
        part = product_id // 1000
        basket = _VOL_BASKET_CACHE.get(vol) or _get_basket_static(vol)
        image_url = f"https://basket-{basket}.wbbasket.ru/vol{vol}/part{part}/{product_id}/images/big/1.webp"

    # ── Данные из поисковой выдачи ──
    position = ''
    is_promo = 'Нет'
    delivery_by = ''
    if search_data:
        position = search_data.get('position', '')
        is_promo = search_data.get('is_promo', 'Нет')
        t1_by = search_data.get('time1_by', '')
        t2_by = search_data.get('time2_by', '')
        if t1_by or t2_by:
            delivery_by = f"{t1_by}-{t2_by} дн." if t1_by and t2_by else f"{t1_by or t2_by} дн."

    # ── Время доставки по Москве ──
    t1_msk = product.get('_time1_msk')
    t2_msk = product.get('_time2_msk')
    delivery_msk = ''
    if t1_msk or t2_msk:
        delivery_msk = f"{t1_msk}-{t2_msk} дн." if t1_msk and t2_msk else f"{t1_msk or t2_msk} дн."

    result = {
        'Артикул': product_id,
        'Бренд': brand,
        'Название': name,
        'Цена без скидки': price_original,
        'Цена со скидкой': price_discounted,
        'Цена с WB кошельком': calc_price_with_wallet(price_discounted),
        'Рейтинг': rating,
        'Отзывы (кол-во)': feedbacks,
        'Остатки всего (шт)': total_stock,
        'Детализация по складам': wh_summary_str,
        'Продавец': supplier,
        'ID Продавца': supplier_id,
        'Позиция в выдаче': position,
        'Реклама (Да/Нет)': is_promo,
        'Срок доставки (Регион)': delivery_by,
        'Срок доставки (МСК)': delivery_msk,
        'Ссылка на товар': f"https://www.wildberries.ru/catalog/{product_id}/detail.aspx",
        'Ссылка на фото': image_url
    }

    return result


# ═══════════════════════════════════════════════════════════════════════════
# WB-КОШЕЛЁК: расчёт цены со скидкой "Незалогиненный кошелёк"
# ═══════════════════════════════════════════════════════════════════════════
#
# Что такое "WB кошелёк"?
#   Wildberries предлагает покупателю-неавторизованному скидку, если он
#   оплатит через кошелёк WB. Скидка задаётся в default-payment.json
#   (процент) и применяется к цене со скидкой.
#
# Как мы это считаем:
#   1) Загружаем default-payment.json → находим "Незалогиненный кошелёк"
#   2) Получаем процент скидки (например, 3%)
#   3) Проверяем лимит по максимальной цене (settings-front.json)
#   4) Применяем: итоговая_цена = цена_со_скидкой * (100 - процент) / 100
#
# Результат кэшируется (= запрашивается один раз за сеанс).
# ───────────────────────────────────────────────────────────────────────────

_WALLET_DISCOUNT: Optional[Decimal] = None
_WALLET_MAX_PRICE: Optional[Decimal] = None


def _fetch_wallet_discount() -> Decimal:
    """Загружает процент скидки для «Незалогиненный кошелёк» из default-payment.json."""
    global _WALLET_DISCOUNT
    if _WALLET_DISCOUNT is not None:
        return _WALLET_DISCOUNT  # уже получили ранее

    url = "https://static-basket-01.wbbasket.ru/vol1/global-payment/default-payment.json"
    try:
        resp = _SESSION.get(url, timeout=10)
        resp.raise_for_status()
        payload = resp.json()
    except Exception:
        return Decimal("0")

    if payload.get("state") != 0:
        return Decimal("0")

    # Ищем в списке тип "Незалогиненный кошелёк" с active=True
    for item in payload.get("data", []):
        if item.get("wc_type") == "Незалогиненный кошелёк" and item.get("is_active") is True:
            try:
                _WALLET_DISCOUNT = Decimal(item["discount_value"])
                return _WALLET_DISCOUNT
            except Exception:
                return Decimal("0")
    return Decimal("0")


def _fetch_wallet_max_price() -> Decimal:
    """
    Загружает максимальную цену, до которой применяется скидка кошелька.
    Если цена товара выше — скидка не действует.
    """
    global _WALLET_MAX_PRICE
    if _WALLET_MAX_PRICE is not None:
        return _WALLET_MAX_PRICE

    url = "https://static-basket-01.wbbasket.ru/vol0/data/settings-front.json"
    try:
        resp = _SESSION.get(url, timeout=10)
        resp.raise_for_status()
        settings = resp.json().get("variables", {})
    except Exception:
        return Decimal("0")

    _WALLET_MAX_PRICE = Decimal(settings.get("wlt1DiscountDisplayMaxPrice", 0))
    return _WALLET_MAX_PRICE


def calc_price_with_wallet(price: float) -> float:
    """
    Применяет скидку WB-кошелька к указанной цене.

    Args:
        price: исходная цена (например, цена со скидкой)

    Returns:
        цена после применения скидки (округляется вниз до целого)
    """
    discount = _fetch_wallet_discount()
    if discount <= 0:
        return price

    max_price = _fetch_wallet_max_price()
    if max_price and Decimal(str(price)) > max_price:
        return price  # цена превышает лимит — скидка не применяется

    # discounted_price = price * (100 - discount%) / 100
    discounted = (Decimal(str(price)) * (Decimal("100") - discount) / Decimal("100")) \
        .quantize(Decimal("1"), rounding=ROUND_FLOOR)
    return float(discounted)


# ═══════════════════════════════════════════════════════════════════════════
# ОСНОВНАЯ ФУНКЦИЯ СБОРА: получение информации по списку артикулов
# ═══════════════════════════════════════════════════════════════════════════

_DETAIL_MIRRORS = [
    "https://card.wb.ru/cards/v4/detail",
    "https://cards.wb.ru/cards/v4/detail",
    "https://card.wb.ru/cards/v5/detail",
    "https://cards.wb.ru/cards/v5/detail",
]

# Регионы для параллельного сбора остатков по складам.
# Чем больше регионов — тем выше шанс получить полную картину stocks[].
# (все dests взяты из анализа TS-парсера, покрывают все основные склады WB)
_DETAIL_DESTS = [
    "123585815",   # Москва / Центр
    "123585590",   # СПб / Северо-Запад
    "123585567",   # Сибирь / Новосибирск
    "123585532",   # Юг / Краснодар
    "123585474",   # Поволжье / Казань
    "123585558",   # Вост. Сибирь / Кемерово
    "-1257786",    # Москва (альт.)
    "-1181704",    # Россия
    "-1029256",    # Россия
    "-1221148",    # Казахстан
    "-2888067",    # Гродно, Беларусь — базовый для цен (BYN)
]

# Базовый регион — из него берём цены (должен поддерживать BYN)
_PRIMARY_DEST = "-2888067"  # Гродно, Беларусь

# Валюта для каждого региона (цены должны быть в правильной валюте,
# иначе WB может вернуть priceU=0)
_DEST_CURRENCIES = {
    "123585815": "rub",
    "123585590": "rub",
    "123585567": "rub",
    "123585532": "rub",
    "123585474": "rub",
    "123585558": "rub",
    "-1257786": "rub",
    "-1181704": "rub",
    "-1029256": "rub",
    "-1221148": "kzt",
    "-2888067": "byn",
}


def _fetch_detail_batch(params: Dict[str, str], proxies: Optional[List[str]] = None, timeout: int = 15) -> Optional[Dict]:
    """
    Пытается выполнить запрос к API деталей товара, перебирая зеркала (mirrors).
    Если одно зеркало недоступно/ошибка — пробует следующее.
    """
    for url in _DETAIL_MIRRORS:
        try:
            response = make_request(url, params=params, headers=DEFAULT_HEADERS, proxies=proxies, timeout=timeout)
            if response.status_code == 200:
                data = response.json()
                if data.get('products') or data.get('data', {}).get('products'):
                    return data
        except Exception:
            continue
    return None

def fetch_products_by_skus(skus: List[int], warehouse_map: Dict[int, str],
                           proxies: Optional[List[str]] = None,
                           search_meta: Optional[Dict[int, Dict[str, Any]]] = None) -> List[Dict[str, Any]]:
    """
    Получает детальную информацию по списку артикулов (SKU).

    Как работает:
      1) Делит артикулы на батчи по 100 шт (ограничение API WB)
      2) Для каждого батча делает параллельные запросы к нескольким регионам
         (каждый — со своей валютой: RUB, BYN, KZT)
      3) Цены берёт из PRIMARY_DEST (-2888067, BYN)
      4) Stocks[] мерджит через Math.max из ВСЕХ регионов
      5) Парсит каждый товар через parse_single_product()

    Зачем несколько регионов?
      - WB для разных dest возвращает разные наборы складов в stocks[]
      - Объединяя их через Math.max, получаем максимально полную картину

    Важно: totalQuantity: 39 — заглушка WB, парсер её игнорирует.
    """
    parsed_products = []

    # API WB принимает максимум 100 артикулов за один запрос
    batch_size = 100
    batches = [skus[i:i + batch_size] for i in range(0, len(skus), batch_size)]

    def _fetch_region(nm, dest, curr):
        params = {
            'appType': '1',
            'curr': curr,
            'dest': dest,
            'spp': '30',
            'nm': nm
        }
        data = _fetch_detail_batch(params, proxies=proxies, timeout=15)
        return dest, (data or {}).get('products') or (data or {}).get('data', {}).get('products', [])

    with Progress(
        SpinnerColumn(),
        TextColumn("[progress.description]{task.description}"),
        BarColumn(),
        TextColumn("[progress.percentage]{task.percentage:>3.0f}%"),
        TimeRemainingColumn(),
        console=console
    ) as progress:
        task = progress.add_task("[cyan]Сбор данных по артикулам...", total=len(batches))

        for batch in batches:
            nm_param = ";".join(map(str, batch))

            try:
                time.sleep(random.uniform(0.5, 1.5))

                # ── Шаг 1: запрашиваем PRIMARY_DEST отдельно (цены в BYN) ──
                primary_curr = _DEST_CURRENCIES.get(_PRIMARY_DEST, DEFAULT_CURR)
                primary_products = _fetch_region(nm_param, _PRIMARY_DEST, primary_curr)[1]

                # Если PRIMARY_DEST не ответил — пробуем ещё раз (с паузой)
                if not primary_products:
                    time.sleep(random.uniform(1.0, 2.0))
                    primary_products = _fetch_region(nm_param, _PRIMARY_DEST, primary_curr)[1]

                if not primary_products:
                    console.print("[yellow]⚠ PRIMARY_DEST не ответил, пропускаем батч.[/yellow]")
                    progress.advance(task)
                    continue

                # ── Шаг 2: параллельные запросы к ОСТАЛЬНЫМ регионам (только stocks) ──
                other_dests = [d for d in _DETAIL_DESTS if d != _PRIMARY_DEST]
                region_results = {_PRIMARY_DEST: primary_products}
                with ThreadPoolExecutor(max_workers=len(other_dests)) as executor:
                    futures = {}
                    for d in other_dests:
                        curr = _DEST_CURRENCIES.get(d, DEFAULT_CURR)
                        futures[executor.submit(_fetch_region, nm_param, d, curr)] = d
                    for future in as_completed(futures):
                        d = futures[future]
                        try:
                            ret_dest, prods = future.result()
                            if prods:
                                region_results[ret_dest] = prods
                        except Exception:
                            pass

                # Собираем продукты по регионам
                # all_instances[pid] = [(dest, product), ...]
                all_instances = {}
                for dest, prods in region_results.items():
                    for p in prods:
                        pid = p.get('id')
                        if pid is not None:
                            all_instances.setdefault(pid, []).append((dest, p))

                # ── Объединяем stocks[] через Math.max ──
                # Базовый продукт — всегда из PRIMARY_DEST (цены гарантированы)
                for prod_id, instances in all_instances.items():
                    prod = None
                    for dest, inst in instances:
                        if dest == _PRIMARY_DEST:
                            prod = inst
                            break
                    if not prod:
                        continue  # без PRIMARY_DEST цены не гарантированы — пропускаем

                    # Собираем размеры → склады → макс. остаток из ВСЕХ регионов
                    merged_stocks = {}  # {size_name: {wh_id: max_qty}}
                    has_any_stock = False

                    for dest, inst in instances:
                        for sz in (inst.get('sizes') or []):
                            sz_name = sz.get('origName') or sz.get('name') or 'No Size'
                            if sz_name not in merged_stocks:
                                merged_stocks[sz_name] = {}
                            for st in (sz.get('stocks') or []):
                                wh = st.get('wh')
                                qty = st.get('qty', 0)
                                if wh is not None:
                                    prev = merged_stocks[sz_name].get(wh, 0)
                                    merged_stocks[sz_name][wh] = prev if prev > qty else qty
                                    if qty > 0:
                                        has_any_stock = True

                    # Применяем объединённые остатки к базовому продукту
                    for sz in (prod.get('sizes') or []):
                        sz_name = sz.get('origName') or sz.get('name') or 'No Size'
                        if sz_name in merged_stocks:
                            wh_qty_list = merged_stocks[sz_name]
                            sz['stocks'] = [{'wh': wh, 'qty': qty} for wh, qty in wh_qty_list.items()]

                    # Берём время доставки из любого инстанса, где оно есть
                    for dest, inst in instances:
                        if 'time1' in inst or 'time2' in inst:
                            prod['_time1_msk'] = inst.get('time1', '')
                            prod['_time2_msk'] = inst.get('time2', '')
                            break

                    parsed_prod = parse_single_product(prod, warehouse_map,
                                                       search_meta.get(prod_id) if search_meta else None)
                    parsed_products.append(parsed_prod)

            except Exception as e:
                console.print(f"[red]Ошибка при обработке пакета артикулов: {e}[/red]")

            progress.advance(task)

    return parsed_products


# ═══════════════════════════════════════════════════════════════════════════
# ПОЛУЧЕНИЕ ТОКЕНА x_wbaas_token (ключ к внутренним API Wildberries)
# ═══════════════════════════════════════════════════════════════════════════
#
# Проблема: Wildberries проверяет, что запрос пришёл от реального браузера.
#   Он запускает JS-челлендж (проверка canvas, WebGL, и т.д.).
#   Если челлендж пройден — браузер получает cookie x_wbaas_token.
#   Без этого токена внутренние API (поиск, цены) возвращают 498.
#
# Решение: мы получаем токен тремя способами (по приоритету):
#   1) SeleniumBase с undetected-chromedriver — полноценный Chrome, проходит всё
#   2) curl_cffi — эмулирует отпечаток Chrome, может сработать
#   3) HTTP + регекс — выкусывает токен из HTML (если был установлен ранее)
#
# Токен кэшируется в _WBAAS_TOKEN на весь сеанс. Его хватает на ~100+ запросов.
# ───────────────────────────────────────────────────────────────────────────

def get_wbaas_token(proxies: Optional[List[str]] = None) -> Optional[str]:
    """
    Пытается получить x_wbaas_token для доступа к внутренним API WB.

    Приоритет методов:
      1. SeleniumBase + undetected Chrome (самый надёжный)
      2. curl_cffi (быстрый, без браузера)
      3. HTTP + регулярное выражение (крайний случай)

    Returns:
        строка токена или None, если не удалось получить
    """
    global _WBAAS_TOKEN
    if _WBAAS_TOKEN:
        return _WBAAS_TOKEN

    # ── Метод 1: SeleniumBase (реальный браузер Chrome) ──
    def _seleniumbase() -> Optional[str]:
        try:
            from seleniumbase import Driver
        except Exception:
            return None

        try:
            from seleniumbase import config as sb_config
            driver_dir = os.path.join(os.environ.get('LOCALAPPDATA', os.path.expanduser('~')),
                                      'WBParser', 'drivers')
            os.makedirs(driver_dir, exist_ok=True)
            sb_config.settings.NEW_DRIVER_DIR = driver_dir

            proxy_str = None
            if proxies:
                raw_proxy = random.choice(proxies)
                proxy_str = raw_proxy.replace("http://", "").replace("https://", "")

            console.print("[cyan]Получение токена через SeleniumBase...[/cyan]")
            driver = Driver(uc=True, headed=False, headless=True,
                            agent=DEFAULT_HEADERS['User-Agent'], proxy=proxy_str)
            try:
                driver.open("https://www.wildberries.ru/")
                # Ждём появления x_wbaas_token (до 15 секунд)
                for _ in range(15):
                    time.sleep(1.0)
                    cookies = driver.execute_cdp_cmd("Network.getAllCookies", {})
                    for cookie in cookies.get("cookies", []):
                        if cookie.get("name") == "x_wbaas_token":
                            token = cookie.get("value")
                            console.print("[green]Сессионный токен получен через браузер.[/green]")
                            return token
            finally:
                driver.quit()  # закрываем браузер
        except Exception:
            pass
        return None

    # Пробуем SeleniumBase
    token = _seleniumbase()
    if token:
        _WBAAS_TOKEN = token
        return token

    # ── Метод 2: curl_cffi (эмуляция без браузера) ──
    try:
        from curl_cffi import requests as curl_requests

        console.print("[cyan]Получение токена через curl_cffi...[/cyan]")

        proxy_dict = None
        if proxies:
            raw = random.choice(proxies)
            proxy_dict = {"http": raw, "https": raw}

        resp = curl_requests.get(
            "https://www.wildberries.ru/",
            impersonate="chrome131",  # подражаем Chrome 131
            proxies=proxy_dict,
            headers=DEFAULT_HEADERS,
            timeout=15
        )
        # Ищем токен в cookie
        for cookie in resp.cookies:
            if cookie.name == "x_wbaas_token":
                token = cookie.value
                console.print("[green]Сессионный токен получен через curl_cffi (cookie).[/green]")
                _WBAAS_TOKEN = token
                return token
        # Если в cookie нет — ищем в HTML
        import re
        match = re.search(r"x_wbaas_token\s*=\s*'([^']+)'", resp.text)
        if match:
            token = match.group(1)
            console.print("[green]Сессионный токен получен через curl_cffi (регекс).[/green]")
            _WBAAS_TOKEN = token
            return token
        console.print("[yellow]  curl_cffi не нашёл токен (JS-челлендж WB не пропускает без браузера).[/yellow]")
    except ImportError:
        console.print("[yellow]  curl_cffi не установлен.[/yellow]")
    except Exception as e_curl:
        console.print(f"[yellow]  curl_cffi ошибка ({e_curl}).[/yellow]")

    # ── Метод 3: HTTP + регулярное выражение (крайний случай) ──
    try:
        response = make_request(
            "https://www.wildberries.ru/",
            headers=DEFAULT_HEADERS,
            proxies=proxies,
            timeout=15
        )
        import re
        match = re.search(r"x_wbaas_token\s*=\s*'([^']+)'", response.text)
        if match:
            token = match.group(1)
            console.print("[green]Сессионный токен получен через HTTP-запрос.[/green]")
            _WBAAS_TOKEN = token
            return token
    except Exception as e_http:
        console.print(f"[yellow]  HTTP-fallback ошибка ({e_http})[/yellow]")

    return None


# ───────────────────────────────────────────────────────────────────────────
# ВСПОМОГАТЕЛЬНАЯ: определение, является ли товар рекламным
# ───────────────────────────────────────────────────────────────────────────

def _detect_promo(product: Dict[str, Any]) -> str:
    """
    Проверяет, является ли товар в поисковой выдаче рекламным (продвижением).

    WB помечает рекламные товары полем panelPromoId.
    Если оно есть и не пустое — товар на платном месте.

    Returns:
        "Да" — реклама, "Нет" — обычный товар
    """
    panel_promo = product.get('panelPromoId')
    if panel_promo and panel_promo not in (0, '0', None):
        return "Да"
    return "Нет"


# ═══════════════════════════════════════════════════════════════════════════
# ПОИСК ТОВАРОВ ПО КЛЮЧЕВОМУ СЛОВУ (через внутренний search API)
# ═══════════════════════════════════════════════════════════════════════════
#
# WB использует внутренний эндпоинт для поиска. Он требует:
#   - x_wbaas_token в cookie
#   - Определённые заголовки (X-Spa-Version, X-Userid)
#   - Параметры региона (dest), сортировки (sort), валюты (curr) и т.д.
#
# Ответ содержит список товаров с их ID, временем доставки и меткой рекламы.
# После сбора ID, функция вызывает fetch_products_by_skus() для деталей.
# ───────────────────────────────────────────────────────────────────────────

def search_products_by_query(query: str, limit_pages: int,
                             warehouse_map: Dict[int, str],
                             proxies: Optional[List[str]] = None) -> List[Dict[str, Any]]:
    """
    Ищет товары по поисковому запросу и собирает детальную информацию.

    Алгоритм:
      1) Получает x_wbaas_token (если ещё нет)
      2) Перебирает страницы поиска (limit_pages), собирает ID товаров
      3) Для каждого товара сохраняет: позицию, рекламный статус, время доставки
      4) Передаёт все ID в fetch_products_by_skus() для деталей

    Args:
        query: поисковый запрос (например, "джинсы женские")
        limit_pages: сколько страниц обработать (~100 товаров на страницу)
        warehouse_map: справочник складов
        proxies: список прокси

    Returns:
        список товаров с детальной информацией
    """
    global _WBAAS_TOKEN
    token = get_wbaas_token(proxies)
    if token:
        _WBAAS_TOKEN = token

    # ── URL внутреннего поискового API WB ──
    url = "https://www.wildberries.ru/__internal/u-search/exactmatch/sng/common/v18/search"

    # ── Заголовки, которые требует этот API ──
    headers = DEFAULT_HEADERS.copy()
    headers.update({
        'Accept': '*/*',
        'Accept-Language': 'ru-RU,ru;q=0.9',
        'X-Requested-With': 'XMLHttpRequest',  # говорит WB, что это AJAX-запрос
        'X-Spa-Version': '13.15.1',             # версия SPA-фронтенда WB
        'X-Userid': '0',                         # 0 = анонимный пользователь
        'Sec-Fetch-Site': 'same-origin',
        'Sec-Fetch-Mode': 'cors',
        'Sec-Fetch-Dest': 'empty',
    })

    import urllib.parse
    encoded_query = urllib.parse.quote(query)
    headers['Referer'] = f"https://www.wildberries.ru/catalog/0/search.aspx?search={encoded_query}"

    # ── Cookie с токеном ──
    cookies = {}
    if token:
        cookies['x_wbaas_token'] = token

    skus_to_fetch = []      # список найденных артикулов
    search_meta = {}        # метаданные: {артикул: {position, is_promo, time1, time2}}
    global_pos = 0          # сквозной счётчик позиции (через все страницы)

    with Progress(
        SpinnerColumn(),
        TextColumn("[progress.description]{task.description}"),
        BarColumn(),
        TextColumn("[progress.percentage]{task.percentage:>3.0f}%"),
        TimeRemainingColumn(),
        console=console
    ) as progress:
        task = progress.add_task(f"[cyan]Поиск артикулов по запросу '{query}'...", total=limit_pages)

        for page in range(1, limit_pages + 1):
            # ── Параметры запроса (как в браузере) ──
            params = [
                ('ab_testing', 'false'),
                ('ab_testing', 'false'),
                ('appType', '64'),
                ('curr', DEFAULT_CURR),
                ('dest', '-2888068'),      # Belarus (для поиска)
                ('hide_dflags', '131072'),
                ('hide_dtype', '11;13'),
                ('hide_vflags', '4294967296'),
                ('inheritFilters', 'false'),
                ('lang', 'ru'),
                ('locale', 'by'),
                ('mdg', '110'),
                ('query', query),
                ('resultset', 'catalog'),
                ('sort', 'popular'),
                ('spp', '30'),
                ('suppressSpellcheck', 'false'),
                ('uclusters', '2'),
            ]
            if page > 1:
                params.append(('page', str(page - 1)))
                params.append(('limit', '300'))

            try:
                time.sleep(random.uniform(0.5, 1.5))
                response = make_request(url, params=params, headers=headers,
                                        cookies=cookies, proxies=proxies, timeout=15)
                data = response.json()

                products_list = data.get('products') or data.get('data', {}).get('products', [])
                if not products_list:
                    console.print(f"[yellow]Страница {page}: Товары закончились или не найдены.[/yellow]")
                    progress.advance(task, advance=limit_pages - page + 1)
                    break

                for i, prod in enumerate(products_list):
                    prod_id = prod.get('id')
                    # Избегаем дубликатов (один товар может быть на нескольких страницах)
                    if prod_id and prod_id not in search_meta:
                        global_pos += 1
                        is_promo = _detect_promo(prod)
                        time1 = prod.get('time1', '')
                        time2 = prod.get('time2', '')
                        search_meta[prod_id] = {
                            'position': global_pos,
                            'is_promo': is_promo,
                            'time1_by': time1,
                            'time2_by': time2
                        }
                        skus_to_fetch.append(prod_id)

            except Exception as e:
                console.print(f"[red]Ошибка при парсинге страницы {page}: {e}[/red]")

            progress.advance(task)

    if not skus_to_fetch:
        console.print("[yellow]Не найдено ни одного артикула по запросу.[/yellow]")
        return []

    console.print(f"[cyan]Найдено артикулов по поиску: {len(skus_to_fetch)}. "
                   f"Загружаем полную детальную информацию о товарах (склады, цены, бренды)...[/cyan]")

    return fetch_products_by_skus(skus_to_fetch, warehouse_map, proxies, search_meta)


# ═══════════════════════════════════════════════════════════════════════════
# ЗАГРУЗКА ПРОКСИ-СЕРВЕРОВ ИЗ ФАЙЛА
# ═══════════════════════════════════════════════════════════════════════════

def load_proxies(filepath: str) -> List[str]:
    """
    Загружает список прокси-серверов из текстового файла.

    Формат файла (каждая строка — один прокси):
      ip:port
      http://ip:port
      http://user:password@ip:port

    Args:
        filepath: путь к файлу

    Returns:
        список строк вида "http://ip:port"
    """
    proxies = []
    if not os.path.exists(filepath):
        console.print(f"[red]Файл с прокси {filepath} не найден.[/red]")
        return proxies

    with open(filepath, 'r', encoding='utf-8') as f:
        for line in f:
            line = line.strip()
            if not line or line.startswith('#'):
                continue
            # Добавляем http:// если не указано
            if not line.startswith('http://') and not line.startswith('https://'):
                proxies.append(f"http://{line}")
            else:
                proxies.append(line)

    console.print(f"[green]Успешно загружено {len(proxies)} прокси.[/green]")
    return proxies


# ═══════════════════════════════════════════════════════════════════════════
# ОТОБРАЖЕНИЕ РЕЗУЛЬТАТОВ В КРАСИВОЙ ТАБЛИЦЕ (через Rich)
# ═══════════════════════════════════════════════════════════════════════════

def display_results_in_table(products: List[Dict[str, Any]], limit: int = 10) -> None:
    """
    Показывает первые N товаров в цветной таблице в консоли.
    Использует библиотеку Rich для красивого форматирования.

    Args:
        products: список товаров
        limit: сколько товаров показать (по умолчанию 10)
    """
    if not products:
        console.print("[yellow]Нет данных для отображения.[/yellow]")
        return

    table = Table(
        title=f"Результаты парсинга (Показано первые {min(limit, len(products))} "
              f"из {len(products)} товаров)",
        box=box.DOUBLE_EDGE,
        header_style="bold magenta",
        title_style="bold cyan"
    )

    table.add_column("Артикул", style="dim", width=12)
    table.add_column("Бренд", style="green", width=15)
    table.add_column("Название", style="white", max_width=30, overflow="ellipsis")
    table.add_column("Цена (без/со ск.)", justify="right", style="cyan")
    table.add_column("Цена (кошелёк)", justify="right", style="green")
    table.add_column("Рейт./Отзывы", justify="center")
    table.add_column("Всего в наличии", justify="right", style="yellow")
    table.add_column("Продавец", style="magenta", max_width=20, overflow="ellipsis")

    for prod in products[:limit]:
        price_str = f"{prod.get('Цена без скидки', 0):.2f} / {prod.get('Цена со скидкой', 0):.2f} BYN"
        wallet_str = f"{prod.get('Цена с WB кошельком', 0):.2f} BYN"
        rating_str = f"★{prod.get('Рейтинг', 0)} ({prod.get('Отзывы (кол-во)', 0)})"
        table.add_row(
            str(prod.get('Артикул', '')),
            str(prod.get('Бренд', '')),
            str(prod.get('Название', '')),
            price_str,
            wallet_str,
            rating_str,
            f"{prod.get('Остатки всего (шт)', 0)} шт",
            str(prod.get('Продавец', ''))
        )

    console.print(table)


# ═══════════════════════════════════════════════════════════════════════════
# ПОЛУЧЕНИЕ ХАРАКТЕРИСТИК И ОПИСАНИЯ ТОВАРА
# ═══════════════════════════════════════════════════════════════════════════
#
# Характеристики (цвет, размер, материал и т.д.) и описание товара
# хранятся в JSON на CDN WB: basket-XX.wbbasket.ru/vol{vol}/.../card.json
#
# Если CDN недоступен — пробуем через v5 и v4 API (медленнее).
# ───────────────────────────────────────────────────────────────────────────

def fetch_product_details(nm_id: int, proxies: Optional[List[str]] = None) -> Dict[str, Any]:
    """
    Загружает описание и характеристики товара.

    Args:
        nm_id: артикул товара
        proxies: список прокси

    Returns:
        словарь: {'description': str, 'characteristics': [{'Характеристика': ..., 'Значение': ...}]}
    """
    vol = nm_id // 100000
    part = nm_id // 1000
    basket = get_basket_dynamically(vol, part, nm_id)

    # ── 1) Пробуем CDN (быстрее всего) ──
    url = f"https://basket-{basket}.wbbasket.ru/vol{vol}/part{part}/{nm_id}/info/ru/card.json"

    try:
        time.sleep(random.uniform(0.1, 0.3))
        response = make_request(url, proxies=proxies, timeout=15)
        if response.status_code == 200:
            product = response.json()
            description = product.get('description', '') or ''
            raw_chars = product.get('options', []) or []
            characteristics = []

            for item in raw_chars:
                name = item.get('name', '')
                value = item.get('value', '')
                if name:
                    characteristics.append({'Характеристика': name, 'Значение': str(value)})

            # grouped_options — это характеристики, сгруппированные по категориям
            for group in product.get('grouped_options', []) or []:
                group_name = group.get('name', '')
                for opt in group.get('options', []):
                    opt_name = opt.get('name', '')
                    opt_value = opt.get('value', '')
                    char_name = f"{group_name} / {opt_name}" if group_name else opt_name
                    if opt_name:
                        characteristics.append({'Характеристика': char_name, 'Значение': str(opt_value)})

            return {'description': description, 'characteristics': characteristics}
    except Exception:
        pass

    # ── 2) Fallback: v5/v4 API (если CDN не ответил) ──
    global _WBAAS_TOKEN
    if not _WBAAS_TOKEN:
        _WBAAS_TOKEN = get_wbaas_token(proxies)
    token_cookies = {'x_wbaas_token': _WBAAS_TOKEN} if _WBAAS_TOKEN else {}

    params = {
        'appType': '1',
        'curr': DEFAULT_CURR,
        'dest': '-2888068',
        'spp': '30',
        'nm': str(nm_id)
    }

    for mirror_url in _DETAIL_MIRRORS:
        if "v4/detail" in mirror_url or "v5/detail" in mirror_url:
            try:
                time.sleep(random.uniform(0.3, 0.8))
                response = make_request(mirror_url, params=params, headers=DEFAULT_HEADERS,
                                         proxies=proxies, timeout=15, cookies=token_cookies)
                if response.status_code == 200:
                    data = response.json()
                    products_list = data.get('data', {}).get('products', []) or data.get('products', [])
                    if products_list:
                        product = products_list[0]
                        description = product.get('description', '') or ''
                        raw_chars = product.get('options', []) or product.get('characteristics', []) or []
                        characteristics = []
                        for item in raw_chars:
                            name = item.get('name', '')
                            value = item.get('value', '')
                            if name:
                                characteristics.append({'Характеристика': name, 'Значение': str(value)})
                        for group in product.get('grouped_options', []) or []:
                            group_name = group.get('name', '')
                            for opt in group.get('options', []):
                                opt_name = opt.get('name', '')
                                opt_value = opt.get('value', '')
                                char_name = f"{group_name} / {opt_name}" if group_name else opt_name
                                if opt_name:
                                    characteristics.append({'Характеристика': char_name, 'Значение': str(opt_value)})
                        return {'description': description, 'characteristics': characteristics}
            except Exception:
                continue

    console.print(f"[yellow]  Предупреждение: Не удалось получить детали товара {nm_id} "
                   f"(все попытки провалились).[/yellow]")
    return {'description': '', 'characteristics': []}


# ───────────────────────────────────────────────────────────────────────────
# ПАРАЛЛЕЛЬНАЯ ЗАГРУЗКА ХАРАКТЕРИСТИК (через ThreadPoolExecutor)
# ───────────────────────────────────────────────────────────────────────────

def fetch_all_details_parallel(products: List[Dict[str, Any]],
                               proxies: Optional[List[str]] = None,
                               max_workers: int = 10) -> Dict[int, Dict]:
    """
    Загружает характеристики товаров параллельно (до 10 потоков),
    что значительно ускоряет процесс для большого списка.

    Args:
        products: список товаров
        proxies: список прокси
        max_workers: количество потоков (по умолчанию 10)

    Returns:
        словарь {артикул: {description, characteristics}}
    """
    details_map = {}
    with Progress(
        SpinnerColumn(),
        TextColumn("[progress.description]{task.description}"),
        BarColumn(),
        TextColumn("[progress.percentage]{task.percentage:>3.0f}%"),
        TimeRemainingColumn(),
        console=console
    ) as progress:
        task = progress.add_task("[cyan]Загрузка характеристик...", total=len(products))

        with ThreadPoolExecutor(max_workers=max_workers) as executor:
            futures = {
                executor.submit(fetch_product_details, p.get('Артикул'), proxies): p
                for p in products if p.get('Артикул')
            }
            for future in as_completed(futures):
                p = futures[future]
                nm_id = p.get('Артикул')
                try:
                    details_map[nm_id] = future.result()
                except Exception:
                    details_map[nm_id] = {'description': '', 'characteristics': []}
                progress.advance(task)

    return details_map


# ═══════════════════════════════════════════════════════════════════════════
# ЭКСПОРТ РЕЗУЛЬТАТОВ В EXCEL И CSV
# ═══════════════════════════════════════════════════════════════════════════

def export_data(
    products: List[Dict[str, Any]],
    base_filename: str,
    include_details: bool = False,
    proxies: Optional[List[str]] = None,
    save_csv: bool = False
) -> None:
    """
    Сохраняет результаты в файлы Excel (.xlsx) и опционально CSV.

    Если include_details=True:
      - В Excel создаётся лист "Товары" с основной таблицей
      - Для каждого товара создаётся отдельный лист (1, 2, 3...) с:
        * артикулом, названием, брендом, продавцом
        * описанием
        * характеристиками
        * фотографией товара (вставлена в лист)

    Args:
        products: список словарей с данными
        base_filename: базовое имя файла (без расширения)
        include_details: загружать ли характеристики/описание
        proxies: список прокси
        save_csv: сохранять ли также CSV-файл
    """
    if not products:
        console.print("[yellow]Экспорт отменен: нет данных для сохранения.[/yellow]")
        return

    df = pd.DataFrame(products)

    excel_file = f"{base_filename}.xlsx"
    csv_file = f"{base_filename}.csv"

    # Если нужно — предзагружаем характеристики
    details_map = {}
    if include_details:
        console.print(f"\n[cyan]Загрузка характеристик и описания для {len(products)} товаров...[/cyan]")
        details_map = fetch_all_details_parallel(products, proxies)

    try:
        with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Товары', index=False)

            ws = writer.sheets['Товары']
            ws.auto_filter.ref = ws.dimensions
            for col_cells in ws.columns:
                max_len = 0
                col_letter = col_cells[0].column_letter
                for cell in col_cells:
                    val = str(cell.value or '')
                    max_len = max(max_len, len(val))
                ws.column_dimensions[col_letter].width = min(max_len + 3, 60)

            if include_details:
                wb = writer.book
                # Удаляем пустой лист по умолчанию (openpyxl создаёт 'Sheet')
                if 'Sheet' in wb.sheetnames:
                    del wb['Sheet']

                for idx, prod in enumerate(products, start=1):
                    nm_id = prod.get('Артикул')
                    sheet_name = str(idx)

                    details = details_map.get(nm_id, {'description': '', 'characteristics': []})
                    chars = details.get('characteristics', [])
                    desc = details.get('description', '')

                    # Создаём лист через openpyxl (шаблон: Поле+Значение слева,
                    # характеристики справа)
                    ws = wb.create_sheet(title=sheet_name)
                    bold_font = Font(name='Calibri', size=11, bold=True)
                    normal_font = Font(name='Calibri', size=11, bold=False)

                    # ── Row 1: заголовки левой панели (центр, bold) ──
                    ws['C1'] = 'Поле'
                    ws['D1'] = 'Значение'
                    ws['C1'].font = bold_font
                    ws['D1'].font = bold_font
                    ws['C1'].alignment = Alignment(horizontal='center', vertical='top')
                    ws['D1'].alignment = Alignment(horizontal='center', vertical='top')

                    # ── Left panel: основные поля товара (колонки C-D) ──
                    info_rows = [
                        ('Артикул', str(nm_id or '')),
                        ('Название', str(prod.get('Название', ''))),
                        ('Бренд', str(prod.get('Бренд', ''))),
                        ('Продавец', str(prod.get('Продавец', ''))),
                        ('Ссылка', str(prod.get('Ссылка на товар', ''))),
                        ('Итого Складских запасов', f"{prod.get('Остатки всего (шт)', 0)} шт"),
                        ('Детализация складов', str(prod.get('Детализация по складам', ''))),
                    ]

                    # ── Right panel: характеристики рядом с левой панелью ──
                    # row 2 = шапка "--- ХАРАКТЕРИСТИКИ ---", row 3+ = значения
                    if chars:
                        cell = ws.cell(row=2, column=6, value='--- ХАРАКТЕРИСТИКИ ---')
                    else:
                        ws.cell(row=2, column=6, value='Характеристики').font = bold_font
                        ws.cell(row=2, column=7, value='Не найдены').font = normal_font

                    for i, (field, value) in enumerate(info_rows):
                        row = i + 2
                        ws.cell(row=row, column=3, value=field).font = bold_font
                        ws.cell(row=row, column=4, value=value).font = normal_font
                        char_idx = i - 1  # row 2 = заголовок, chars[0] = row 3
                        if 0 <= char_idx < len(chars):
                            ws.cell(row=row, column=6, value=chars[char_idx].get('Характеристика', '')).font = bold_font
                            ws.cell(row=row, column=7, value=chars[char_idx].get('Значение', '')).font = normal_font

                    # Оставшиеся характеристики (после 6, которые влезли рядом)
                    if chars and len(chars) > len(info_rows) - 1:
                        for ci in range(len(info_rows) - 1, len(chars)):
                            row = ci + 3
                            ws.cell(row=row, column=6, value=chars[ci].get('Характеристика', '')).font = bold_font
                            ws.cell(row=row, column=7, value=chars[ci].get('Значение', '')).font = normal_font

                    # ── Описание внизу (после пустой строки) ──
                    last_content_row = max(len(info_rows) + 1, len(chars) + 2)
                    desc_label_row = last_content_row + 2  # blank row gap
                    if desc:
                        ws.cell(row=desc_label_row, column=3, value='ОПИСАНИЕ').font = bold_font
                        ws.cell(row=desc_label_row, column=3).alignment = Alignment(vertical='top')
                        ws.cell(row=desc_label_row, column=4, value=desc).font = normal_font
                        ws.cell(row=desc_label_row, column=4).alignment = Alignment(
                            horizontal='left', vertical='top', wrap_text=True)
                        ws.row_dimensions[desc_label_row].height = 300

                    # ── Ширина колонок (как в шаблоне) ──
                    ws.column_dimensions['A'].width = 12
                    ws.column_dimensions['B'].width = 37.7
                    ws.column_dimensions['C'].width = 25.29
                    ws.column_dimensions['D'].width = 98
                    ws.column_dimensions['F'].width = 28.86
                    ws.column_dimensions['G'].width = 46.14

                    # ── Фото товара ──
                    if nm_id:
                        vol = nm_id // 100000
                        part = nm_id // 1000
                        basket = get_basket_dynamically(vol, part, nm_id)
                        image_url = f"https://basket-{basket}.wbbasket.ru/vol{vol}/part{part}/{nm_id}/images/big/1.webp"
                    else:
                        image_url = prod.get('Ссылка на фото')

                    if image_url:
                        try:
                            import io
                            from openpyxl.drawing.image import Image as OpenpyxlImage

                            img_resp = _SESSION.get(image_url, timeout=5)
                            if img_resp.status_code == 200:
                                img_data = io.BytesIO(img_resp.content)
                                img = OpenpyxlImage(img_data)

                                max_height = 300
                                if img.height > max_height:
                                    ratio = max_height / img.height
                                    img.height = max_height
                                    img.width = int(img.width * ratio)

                                ws.add_image(img, 'A1')
                        except Exception:
                            pass

        console.print(f"[green]✔ Данные успешно экспортированы в Excel: {excel_file}[/green]")
        if include_details:
            console.print(f"[green]  Дополнительно: создано {len(products)} листов с характеристиками "
                           f"(листы 1–{len(products)}).[/green]")

    except Exception as e:
        console.print(f"[red]Критическая ошибка при экспорте: {e}[/red]")
        # Аварийное сохранение в CSV (без характеристик, только основная таблица)
        emergency_file = f"{base_filename}_emergency.csv"
        try:
            df.to_csv(emergency_file, index=False, encoding='utf-8-sig')
            console.print(f"[yellow]⚠ Аварийное сохранение: {emergency_file} "
                           f"({len(products)} товаров)[/yellow]")
        except Exception:
            console.print("[red]✘ Не удалось выполнить аварийное сохранение.[/red]")
        raise

    # Сохраняем CSV (опционально)
    if save_csv:
        try:
            df.to_csv(csv_file, index=False, encoding='utf-8-sig')
            console.print(f"[green]✔ Данные успешно экспортированы в CSV: {csv_file}[/green]")
        except Exception as e:
            console.print(f"[red]Ошибка при записи CSV-файла: {e}[/red]")


# ═══════════════════════════════════════════════════════════════════════════
# ГЛАВНАЯ ФУНКЦИЯ: точка входа в программу
# ═══════════════════════════════════════════════════════════════════════════
#
# main() — это "дирижёр" всего парсера. Она:
#   1) Разбирает аргументы командной строки
#   2) Спрашивает у пользователя, что делать (если не указано в аргументах)
#   3) Вызывает нужные функции сбора данных
#   4) Показывает результаты и предлагает экспортировать
# ───────────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Парсер товаров Wildberries на Python")
    parser.add_argument("-m", "--mode", choices=["sku", "search"],
                        help="Режим работы: 'sku' (по списку артикулов) или 'search' (по ключевому слову)")
    parser.add_argument("-s", "--skus",
                        help="Список артикулов через запятую (для режима sku)")
    parser.add_argument("-q", "--query",
                        help="Поисковый запрос (для режима search)")
    parser.add_argument("-p", "--pages", type=int, default=5,
                        help="Количество страниц для поиска (по умолчанию 5)")
    parser.add_argument("--proxy-file",
                        help="Путь к файлу со списком прокси")
    parser.add_argument("-o", "--output", default="wb_results",
                        help="Базовое имя файла для сохранения результатов (без расширения)")
    parser.add_argument("--supplier-id", type=int,
                        help="ID продавца (поставщика) для фильтрации результатов")

    args = parser.parse_args()

    # Приветствие
    console.print(Panel.fit(
        "[bold cyan]Парсер товаров Wildberries (WB)[/bold cyan]\n"
        "[dim]Инструмент для сбора базовой и расширенной информации о товарах[/dim]",
        border_style="cyan"
    ))

    # ── Загрузка прокси ──
    proxies = None
    if args.proxy_file:
        proxies = load_proxies(args.proxy_file)
    elif os.path.exists("proxies.txt"):
        proxies = load_proxies("proxies.txt")

    # ── Загрузка справочника складов ──
    console.print("[cyan]Загрузка справочника складов WB для декодирования названий складов...[/cyan]")
    warehouse_map = fetch_warehouse_map(proxies)
    if warehouse_map:
        console.print(f"[green]Справочник складов успешно загружен ({len(warehouse_map)} записей).[/green]")

    mode = args.mode

    # ── Если режим не указан в аргументах — спрашиваем ──
    if not mode:
        console.rule("[bold yellow]Выбор режима работы[/bold yellow]")
        console.print("[bold cyan]1[/bold cyan] — Поиск товаров по ключевому слову")
        console.print("[bold cyan]2[/bold cyan] — Сбор данных по артикулам (SKU)")
        mode_choice = Prompt.ask(
            "Выберите режим (введите 1 или 2)",
            choices=["1", "2"],
            default="1",
            show_choices=False
        )
        if mode_choice == "2":
            mode = "sku"
            console.print("[cyan]Выбран режим: Сбор данных по артикулам (SKU)[/cyan]")
        else:
            mode = "search"
            console.print("[cyan]Выбран режим: Поиск товаров по ключевому слову[/cyan]")

    products_data = []

    # ── Режим 1: по артикулам ──
    if mode == "sku":
        skus_str = args.skus
        if not skus_str:
            skus_str = Prompt.ask("Введите артикулы товаров (через запятую или пробел)")

        # Разбираем строку с артикулами (могут быть через запятую или пробел)
        skus_list = []
        for val in skus_str.replace(',', ' ').split():
            try:
                skus_list.append(int(val.strip()))
            except ValueError:
                continue  # если попалось не число — пропускаем

        if not skus_list:
            console.print("[red]Ошибка: Не указано ни одного корректного артикула.[/red]")
            return

        console.print(f"[cyan]Начинаем сбор по {len(skus_list)} артикулам...[/cyan]")
        products_data = fetch_products_by_skus(skus_list, warehouse_map, proxies)

    # ── Режим 2: по поисковому запросу ──
    elif mode == "search":
        query = args.query
        if not query:
            query = Prompt.ask("Введите ключевое слово для поиска")

        pages = args.pages
        if not args.pages or args.pages == 5:
            pages_str = Prompt.ask(
                "Укажите количество страниц для парсинга (1 страница ~ 100 товаров)",
                default="3"
            )
            try:
                pages = int(pages_str)
            except ValueError:
                pages = 3

        console.print(f"[cyan]Запуск поиска по запросу: '{query}' ({pages} стр.)...[/cyan]")
        products_data = search_products_by_query(query, pages, warehouse_map, proxies)

    # ── Обработка результатов ──
    if products_data:
        console.print(f"\n[bold green]Успешно собрано товаров: {len(products_data)}[/bold green]")

        # Фильтрация по ID продавца
        supplier_id = args.supplier_id
        if not supplier_id:
            if Confirm.ask("Хотите сделать выборку по ID продавца (поставщика)?", default=False):
                supplier_id_str = Prompt.ask("Введите ID продавца (например, 682757)")
                try:
                    supplier_id = int(supplier_id_str.strip())
                except ValueError:
                    console.print("[red]Некорректный ID продавца. Фильтрация отменена.[/red]")

        if supplier_id:
            filtered_data = [p for p in products_data if p.get('ID Продавца') == supplier_id]
            console.print(f"[cyan]Применена выборка по ID продавца {supplier_id}. "
                           f"Было товаров: {len(products_data)}, Стало: {len(filtered_data)}.[/cyan]")
            products_data = filtered_data

        if not products_data:
            console.print("[yellow]В результате выборки не осталось ни одного товара. Экспорт отменен.[/yellow]")
            return

        # Показываем первые 10 товаров в красивой таблице
        display_results_in_table(products_data, limit=10)

        # Спрашиваем, нужна ли выгрузка характеристик
        console.rule("[bold yellow]Выгрузка характеристик и описания[/bold yellow]")
        console.print("[bold cyan]1[/bold cyan] — Да, выгрузить Характеристики и Описание для каждого товара "
                       "(отдельный лист в Excel)")
        console.print("[bold cyan]2[/bold cyan] — Нет, только основные данные")
        details_choice = Prompt.ask(
            "Нужна ли выгрузка Характеристик и описания по выбранным товарам?",
            choices=["1", "2"],
            default="2",
            show_choices=False
        )
        include_details = (details_choice == "1")

        save_csv = Confirm.ask("Сохранить также CSV-файл?", default=False)

        # Авто-имя файла
        today = datetime.now().strftime("%d%m%Y")
        default_name = f"Rezalt_{today}" if include_details else f"RezaltAll_{today}"
        output_name = args.output
        if not args.output or args.output == "wb_results":
            output_name = Prompt.ask(
                "Введите имя файла для сохранения результатов",
                default=default_name
            )

        export_data(products_data, output_name, include_details=include_details, proxies=proxies, save_csv=save_csv)

        csv_msg = f"\n - [cyan]{output_name}.csv[/cyan] (формат CSV UTF-8)" if save_csv else ""
        extra_sheets_msg = (
            f"\n - [cyan]{output_name}.xlsx[/cyan] содержит листы 1–{len(products_data)} "
            f"с характеристиками" if include_details else ""
        )
        console.print(Panel(
            "[bold green]Парсинг успешно завершен![/bold green]\n"
            f"Созданы файлы:\n"
            f" - [cyan]{output_name}.xlsx[/cyan] (лист 'Товары' — основные данные){extra_sheets_msg}{csv_msg}\n"
            "Вы можете открыть их для дальнейшего анализа.",
            title="Результат",
            border_style="green"
        ))
    else:
        console.print("[red]К сожалению, не удалось собрать данные. "
                       "Проверьте подключение к сети или настройки прокси.[/red]")


# ═══════════════════════════════════════════════════════════════════════════
# ТОЧКА ВХОДА
# ═══════════════════════════════════════════════════════════════════════════

if __name__ == '__main__':
    while True:
        try:
            main()
        except KeyboardInterrupt:
            console.print("\n[yellow]Работа парсера прервана пользователем.[/yellow]")
        console.print("\nНажмите Enter для продолжения работы или Esc для выхода из программы")
        key = msvcrt.getch()
        if key == b'\x1b':  # Esc
            break
        # Enter (b'\r') — продолжаем цикл
